import tkinter as tk
from tkinter import messagebox, ttk
from tkinter import filedialog
from openpyxl import workbook, worksheet,load_workbook
from datetime import datetime
import xlsxwriter
from openpyxl.utils import coordinate_from_string, column_index_from_string
from copy import *

"""
prognoza:6464FE
dane pewne 90EE90
niepewne C70000
reczne EF0095
brak schem 64FEFE
brak danych C7C7C7

data FFFFA500
"""
class MainWindow():

    def __init__(self,master):
        self.master=master
        self.master.title("Nagrody za kody")
        self.master.geometry('700x500')

        self.frame1=tk.Frame(master)
        self.frame1.pack(fill='both', expand=True)

        self.button_import=tk.Button(self.frame1, text="Import plików Excel", command=lambda: self.file_choose())
        self.button_import.pack()

        self.button_save=tk.Button(self.frame1, text="Podaj lokalizację zapisu", command=lambda: self.choose_save_location())
        self.button_save.pack()

        self.label_file_count=tk.Label(self.frame1)
        self.label_file_count.pack()

        self.var_win = tk.IntVar()
        self.checkbox_win=tk.Checkbutton(self.frame1, text="Sprawdzaj max/min dla PPE", variable=self.var_win, command=lambda: self.label_state(self.var_win.get()))
        self.checkbox_win.pack(anchor="e")

        self.win_count=tk.StringVar()
        self.win_entry=tk.Entry(self.frame1, textvariable=self.win_count, state="disabled")
        self.win_entry.pack(anchor="e")

        self.var_SE = tk.IntVar()
        self.checkbox_SE=tk.Checkbutton(self.frame1, text="Nie uwzględniaj sprzedawców kompleksowych", variable=self.var_SE)
        self.checkbox_SE.pack(anchor="e")

        self.var1 = tk.IntVar()
        self.checkbox_1=tk.Checkbutton(self.frame1, text="Dane pewne - kolor zielony", variable=self.var1)
        self.checkbox_1.pack(anchor="w")

        self.var2 = tk.IntVar()
        self.checkbox_2=tk.Checkbutton(self.frame1, text="Dane niepewne - kolor czerwony", variable=self.var2)
        self.checkbox_2.pack(anchor="w")

        self.var3 = tk.IntVar()
        self.checkbox_3=tk.Checkbutton(self.frame1, text="Dane prognozowane - kolor niebieski", variable=self.var3)
        self.checkbox_3.pack(anchor="w")

        self.var4 = tk.IntVar()
        self.checkbox_4=tk.Checkbutton(self.frame1, text="Dane ręczne - kolor różowy", variable=self.var4)
        self.checkbox_4.pack(anchor="w")

        self.var5 = tk.IntVar()
        self.checkbox_5=tk.Checkbutton(self.frame1, text="Braki schematów - kolor turkusowy", variable=self.var5)
        self.checkbox_5.pack(anchor="w")

        self.var6 = tk.IntVar()
        self.checkbox_6=tk.Checkbutton(self.frame1, text="Braki danych - kolor szary", variable=self.var6)
        self.checkbox_6.pack(anchor="w")    

        self.var7 = tk.IntVar()
        self.checkbox_7=tk.Checkbutton(self.frame1, text="Zamknięta umowa - kolor biały", variable=self.var7)
        self.checkbox_7.pack(anchor="w")

        self.start_button=tk.Button(self.frame1, text="Start", command=lambda: self.start(imported_files,location,self.progress_var))
        self.start_button.pack(anchor="s")

      
     

        self.progress_var = tk.DoubleVar()
        self.label_progress_var=tk.StringVar()

    def file_choose(self):
        files_dir = filedialog.askopenfilenames()
        global imported_files
        imported_files = files_dir

    def label_state(self,checkbox_status):
        if checkbox_status == 1:
            self.win_entry.configure(state='normal')
        else:
            self.win_entry.configure(state='disabled')


    def choose_save_location(self):
        global location
        location=filedialog.askdirectory()

# zapis do pliku
    def result_file(self,location,value):
        report_date=datetime.now()
        result_excel = xlsxwriter.Workbook(str(location)+'/'+'_Raport PPE - '+datetime.strftime(report_date,"%Y-%m-%d_%H-%M-%S")+'.xlsx')    
        result_sheet = result_excel.add_worksheet('Wynik')
        
        result_sheet.set_column(0,0,35)
        result_sheet.set_column(1,1,38)
        result_sheet.write('A1', 'Kod PPE')
        result_sheet.write('B1', 'Schemat taryfowy danych zatwierdzonych')

        row=1
        for (n,m) in zip(value[0],value[1]):
            result_sheet.write(row,0,n)
            result_sheet.write(row,1,m)
            row+=1            
        
        if self.var_win.get()==1:
            windsor_sheet = result_excel.add_worksheet('Winsor')
            windsor_sheet.write('A1', 'Kod PPE z Win_max')
            windsor_sheet.write('B1', 'Max value')

            windsor_sheet.write('D1', 'Kod PPE Win_min')
            windsor_sheet.write('E1', 'Min value')
            row=1
            for (n,m,v,b) in zip(value[2],value[3],value[4],value[5]):
                windsor_sheet.write(row,0,n)
                windsor_sheet.write(row,1,m)
                windsor_sheet.write(row,3,v)
                windsor_sheet.write(row,4,b)
                row+=1
        result_excel.close()
    
    def start(self,excel_files,location,current_file):


        self.progress = ttk.Progressbar(self.frame1, orient="horizontal", length=500, mode="determinate", maximum=len(excel_files), variable=current_file)
        self.progress.pack()
        self.progress_label=tk.Label(self.frame1, textvariable=self.label_progress_var)
        self.progress_label.pack()
        

        if self.var_win.get()==1:
            try:
                int(self.win_count.get())
            except (TypeError, ValueError):
                tk.messagebox.showerror("Błąd","Naucz się cyferek i spróbuj raz jeszcze!")
                return
        
        self.result_file(location, self.data_extraction(excel_files))
            
    def cell_info(self, cell):
        a = cell.row
        b = cell.column
        return b, a

    def SE_str_split(self,SE_name):
        try:
            
            SE_split=list(SE_name)

            if SE_name=="Nazwa SE" or type(SE_name)==None or SE_name=="-" or len(SE_name)<15:
                return False
            elif SE_split[12] == '5':
                return True
            else:
                return False
        except TypeError:
            return False

    def change_label_text(self, variable, text):
        variable.set(text)

    def data_extraction(self, excel_files):
        # if self.var_win.get()==1:
        #     try:
        #         int(self.win_count.get())
        #     except (TypeError, ValueError):
        #         tk.messagebox.showerror("Błąd","Naucz się cyferek i spróbuj raz jeszcze!")
            

        
        #checboxes=[(self.var1.get(),'FF90EE90'),(self.var2.get(),'FFC70000'),(self.var3.get(),'FF6464FE'),(self.var4.get(),'FFEF0095'),(self.var5.get(),'FF64FEFE'),(self.var6.get(),'FFC7C7C7')]
        
        

        checboxes=[self.var1.get(),self.var2.get(),self.var3.get(),self.var4.get(),self.var5.get(),self.var6.get(),self.var7.get()]
        hexacodes=['FF90EE90','FFC70000','FF6464FE','FFEF0095','FF64FEFE','FFC7C7C7','00000000']

        values_taken=[]
        tariff_taken=[]
        windsor_taken_max_PPE=[]
        windsor_taken_min_PPE=[]
        windsor_taken_max=[]
        windsor_taken_min=[]

        

        for i in excel_files:
            self.change_label_text(self.label_progress_var,i)
            self.progress_var.set(excel_files.index(i)+1)
            self.progress.update()
            self.progress_label.update()
         
            wb1=load_workbook(i, data_only=True) 
            ws=wb1.active
 
            for row in ws.iter_rows(min_row=2, max_row=10):
                zxc=(cell for cell in row if cell.value is not None)
                for cell in zxc:
                    if cell.value=="Energia":
                        energia_cell=self.cell_info(cell)
                    elif cell.value=="Kod PPE":
                        kod_PPE_cell=self.cell_info(cell)
                    elif cell.value=="Schemat taryfowy danych zatwierdzonych":
                        tariff_scheme_cell=self.cell_info(cell)
                    elif cell.value=="Nazwa SE":
                        SE_name_cell=self.cell_info(cell)

            qwe = (row for row in ws[energia_cell[0]] if row.value is not None)

                    
            for row in qwe:
                windsor_list=[]
                col = column_index_from_string(energia_cell[0])+1
                if row.fill.start_color.index in hexacodes:
                    PPE_number_winds=""
                    while  (ws.cell(row=energia_cell[1], column=col).fill.start_color.index == "FFFFA500"):
                        z=ws.cell(row=row.row, column=col).fill.start_color.index

                        if self.var_SE.get()==0:
                                if z != "FFFFA500":
                                    z_value=ws.cell(row=row.row, column=col).value
                                    if z!= '00000000':
                                        windsor_list.append(z_value)
                                        PPE_number_winds=ws[str(kod_PPE_cell[0])+str(row.row)].value

                                if z in hexacodes:
                                    ind=hexacodes.index(z)
                                    
                                    if checboxes[ind]==1:
                                        PPE_number=ws[str(kod_PPE_cell[0])+str(row.row)].value
                                        try:
                                            tariff_scheme=ws[str(tariff_scheme_cell[0])+str(row.row)].value
                                        except UnboundLocalError:                                    
                                            tariff_scheme=str('brak kolumny taryfy w pliku')

                                        if PPE_number not in values_taken:
                                            values_taken.append(PPE_number)
                                            tariff_taken.append(tariff_scheme)
                                col+=1
                        else:    
                            if self.SE_str_split(ws[str(SE_name_cell[0])+str(row.row)].value)!=True:
                                                                                       
                                if z != "FFFFA500":
                                    z_value=ws.cell(row=row.row, column=col).value
                                    if z!= '00000000':
                                        windsor_list.append(z_value)
                                        PPE_number_winds=ws[str(kod_PPE_cell[0])+str(row.row)].value

                                if z in hexacodes:
                                    ind=hexacodes.index(z)
                                    
                                    if checboxes[ind]==1:
                                        PPE_number=ws[str(kod_PPE_cell[0])+str(row.row)].value
                                        try:
                                            tariff_scheme=ws[str(tariff_scheme_cell[0])+str(row.row)].value
                                        except UnboundLocalError:                                    
                                            tariff_scheme=str('brak kolumny taryfy w pliku')

                                        if PPE_number not in values_taken:
                                            values_taken.append(PPE_number)
                                            tariff_taken.append(tariff_scheme)
                            col+=1

                            
                    if self.var_win.get()==1:
                        try:
                            windsor_list_bis=copy(windsor_list)
                            
                            asdasd=(max(windsor_list_bis),min(windsor_list_bis))
                            windsor_list_bis.remove(asdasd[0])
                            windsor_list_bis.remove(asdasd[1])
                            winds_len=len(windsor_list_bis)
                            wins_sum=0
                            for i in windsor_list_bis:
                                wins_sum+=i
                                
                            try:
                                winds_sr=wins_sum/winds_len
                            except ZeroDivisionError:
                                winds_sr=0

                            i_max_value=1
                            i_min_value=1
                            for i in windsor_list:
                                try:
                                    i_value= i/winds_sr
                                except ZeroDivisionError:
                                    i_value=1
                                
                                if i_value > i_max_value and len(windsor_taken_max_PPE)<int(self.win_count.get()):
                                    i_max_value = i_value
                                    if PPE_number_winds not in windsor_taken_max_PPE:
                                        windsor_taken_max_PPE.append(PPE_number_winds)
                                        windsor_taken_max.append(i_max_value)
                                    else:
                                        i_index=windsor_taken_max_PPE.index(PPE_number_winds)
                                        if windsor_taken_max[i_index]<i_max_value:
                                            windsor_taken_max[i_index]=i_max_value 
                                elif i_value > i_max_value and len(windsor_taken_max_PPE)==int(self.win_count.get()):
                                    i_max_value = i_value
                                    if PPE_number_winds in windsor_taken_max_PPE:
                                        i_index=windsor_taken_max_PPE.index(PPE_number_winds)
                                        if windsor_taken_max[i_index]<i_max_value:
                                            windsor_taken_max[i_index]=i_max_value                                    
                                    else:
                                        i_index=windsor_taken_max.index(min(windsor_taken_max))
                                        if windsor_taken_max[i_index]<i_max_value:
                                            windsor_taken_max[i_index]=i_max_value
                                            windsor_taken_max_PPE[i_index]=PPE_number_winds

                                if i_value < i_min_value and len(windsor_taken_min_PPE)<int(self.win_count.get()):
                                    i_min_value = i_value
                                    if PPE_number_winds not in windsor_taken_min_PPE:
                                        windsor_taken_min_PPE.append(PPE_number_winds)
                                        windsor_taken_min.append(i_min_value)
                                    else:
                                        i_index=windsor_taken_min_PPE.index(PPE_number_winds)
                                        if windsor_taken_min[i_index]>i_min_value:
                                            windsor_taken_min[i_index]=i_min_value 
                                elif i_value < i_min_value and len(windsor_taken_min_PPE)==int(self.win_count.get()):
                                    i_min_value = i_value
                                    if PPE_number_winds in windsor_taken_min_PPE:
                                        i_index=windsor_taken_min_PPE.index(PPE_number_winds)
                                        if windsor_taken_min[i_index]>i_min_value:
                                            windsor_taken_min[i_index]=i_min_value                                    
                                    else:
                                        i_index=windsor_taken_min.index(max(windsor_taken_min))
                                        if windsor_taken_min[i_index]>i_min_value:
                                            windsor_taken_min[i_index]=i_min_value
                                            windsor_taken_min_PPE[i_index]=PPE_number_winds

                        except ValueError:
                            pass
        
        return values_taken, tariff_taken, windsor_taken_max_PPE, windsor_taken_max, windsor_taken_min_PPE, windsor_taken_min

        
root=tk.Tk()
my_gui=MainWindow(root)
root.mainloop()
